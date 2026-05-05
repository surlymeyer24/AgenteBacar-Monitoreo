"""
Genera inventario/src/main/resources/data/camaras-activas.json desde CSV o Excel.

CSV (export típico de Excel con «;»):
  Direccion IP;Port;Nombre Dispositivo;…;manufacturer;nombre camara;tipo
  Codificación habitual en Windows: cp1252; separador «;».

Excel (.xlsx): primera fila = encabezados; se detectan columnas por nombre (no por índice fijo).

Uso:
  pip install openpyxl   # solo si usás .xlsx
  py -3 scripts/excel_to_camaras_json.py [ruta.csv|ruta.xlsx]

Sin argumentos: intenta Libro2.csv en Downloads y luego «camaras activas.xlsx».
"""
from __future__ import annotations

import csv
import json
import re
import sys
import unicodedata
from pathlib import Path


def str_or_none(x):
    if x is None:
        return None
    s = str(x).strip()
    return s if s else None


def int_port(p):
    if p is None:
        return None
    if isinstance(p, float):
        return int(p)
    try:
        return int(str(p).strip())
    except (ValueError, TypeError):
        return None


def normalize_header(text) -> str:
    if text is None:
        return ""
    s = unicodedata.normalize("NFD", str(text))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def row_to_record(ip, port, dispositivo, manufacturer, nombre, tipo) -> dict | None:
    nv = str_or_none(dispositivo)
    if not nv:
        return None
    ip_s = str_or_none(ip)
    port_i = int_port(port)
    nombre_s = str_or_none(nombre) or nv
    marca_s = str_or_none(manufacturer)
    tipo_s = str_or_none(tipo)
    ip_part = ip_s if ip_s else ""
    tipo_part = tipo_s if tipo_s else ""
    desc = f"IP {ip_part} — Puerto {port_i} — Dispositivo: {nv} — {tipo_part}".strip()
    return {
        "nombreDispositivo": nv,
        "nombre": nombre_s,
        "ubicacion": nombre_s,
        "marca": marca_s,
        "descripcion": desc,
        "direccionIp": ip_s,
        "puerto": port_i,
        "tipo": tipo_s,
    }


def read_csv_semicolon(path: Path) -> list[dict]:
    """CSV separado por «;», típico export Excel ES (cp1252)."""
    rows_out = []
    for encoding in ("cp1252", "utf-8-sig", "latin-1"):
        try:
            with open(path, "r", encoding=encoding, newline="") as f:
                reader = csv.reader(f, delimiter=";")
                raw = list(reader)
            break
        except UnicodeDecodeError:
            raw = None
    else:
        raise RuntimeError("No se pudo decodificar el CSV (probá guardar como UTF-8).")

    if not raw:
        return []

    start = 0
    if raw[0] and len(raw[0]) >= 3:
        h2 = normalize_header(raw[0][2])
        if "dispositivo" in h2:
            start = 1

    for row in raw[start:]:
        if not row or len(row) < 7:
            continue
        ip, port, dispositivo = row[0], row[1], row[2]
        manufacturer = row[4] if len(row) > 4 else None
        nombre = row[5] if len(row) > 5 else None
        tipo = row[6] if len(row) > 6 else None
        rec = row_to_record(ip, port, dispositivo, manufacturer, nombre, tipo)
        if rec:
            rows_out.append(rec)
    return rows_out


def read_xlsx(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Para .xlsx: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return []

    col_map: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        key = normalize_header(cell)
        col_map[key] = i

    def pick(*candidates: str) -> int | None:
        for cand in candidates:
            c = normalize_header(cand)
            if c in col_map:
                return col_map[c]
        for stored, idx in col_map.items():
            for cand in candidates:
                if normalize_header(cand) in stored or stored in normalize_header(cand):
                    return idx
        return None

    idx_ip = pick("Direccion IP", "IP", "DireccionIp")
    idx_port = pick("Port", "Puerto")
    idx_dev = pick("Nombre Dispositivo", "Dispositivo")
    idx_manu = pick("manufacturer", "Marca")
    idx_nombre = pick("nombre camara", "Nombre camara", "nombre", "Nombre")
    idx_tipo = pick("tipo", "Tipo", "modelo", "Modelo")

    if idx_dev is None:
        wb.close()
        raise ValueError(
            "En el Excel no encontré la columna «Nombre Dispositivo» / «Dispositivo». "
            f"Encabezados vistos: {list(col_map.keys())}"
        )

    rows_out = []
    for row in rows_iter:
        if not row:
            continue
        def at(j: int | None):
            return row[j] if j is not None and j < len(row) else None

        rec = row_to_record(
            at(idx_ip),
            at(idx_port),
            at(idx_dev),
            at(idx_manu),
            at(idx_nombre),
            at(idx_tipo),
        )
        if rec:
            rows_out.append(rec)

    wb.close()
    return rows_out


def default_input_path() -> Path | None:
    downloads = Path.home() / "Downloads"
    candidates = [
        downloads / "Libro2.csv",
        downloads / "camaras activas.xlsx",
    ]
    for p in candidates:
        if p.is_file():
            return p
    return None


def main():
    inv = Path(__file__).resolve().parents[1]
    out_path = inv / "src/main/resources/data/camaras-activas.json"

    if len(sys.argv) > 1:
        src = Path(sys.argv[1]).expanduser()
    else:
        src = default_input_path()
        if src is None:
            print(
                "Pasá la ruta al CSV o XLSX:\n"
                "  py -3 scripts/excel_to_camaras_json.py C:\\ruta\\Libro2.csv",
                file=sys.stderr,
            )
            sys.exit(1)

    if not src.is_file():
        print(f"No existe: {src}", file=sys.stderr)
        sys.exit(1)

    suf = src.suffix.lower()
    if suf == ".csv":
        rows_out = read_csv_semicolon(src)
    elif suf in (".xlsx", ".xlsm"):
        rows_out = read_xlsx(src)
    else:
        print(f"Formato no soportado: {suf} (usá .csv o .xlsx)", file=sys.stderr)
        sys.exit(1)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(rows_out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"{len(rows_out)} filas desde {src.name} -> {out_path}")


if __name__ == "__main__":
    main()
